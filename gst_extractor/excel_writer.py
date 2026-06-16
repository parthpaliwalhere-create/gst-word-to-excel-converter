from __future__ import annotations

from collections import defaultdict
from datetime import datetime
import re
from pathlib import Path
from typing import List, Optional, Dict

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter

from .extractor import InvoiceRecord, ItemRecord

MONTH_NAMES = ["JAN", "FEB", "MARCH", "APRIL", "MAY", "JUNE", "JULY", "AUG", "SEPT", "OCT", "NOV", "DEC"]
MONTH_FULL = ["JANUARY", "FEBRUARY", "MARCH", "APRIL", "MAY", "JUNE", "JULY", "AUGUST", "SEPTEMBER", "OCTOBER", "NOVEMBER", "DECEMBER"]

HEADERS = ["DATE", "BILL NO", "NAME", "GSTIN", "HSN", "TAXABLE", "IGST", "CGST", "SGST", "TOTAL", "RATE"]

NAVY = "17365D"
HEADER_BLUE = "D9EAF7"
LIGHT_BLUE = "EAF3F8"
YELLOW = "FFF2CC"
RED = "F8CBAD"
GREEN = "E2F0D9"
WHITE = "FFFFFF"
BORDER_COLOR = "7F7F7F"
GREY = "F2F2F2"
SPECIAL_ORANGE = "FFC000"
SPECIAL_RED = "9C0006"



def _safe_sheet_name(name: str) -> str:
    for ch in '[]:*?/\\':
        name = name.replace(ch, "-")
    return name[:31]


def _parse_date(date_text: str) -> Optional[datetime]:
    if not date_text:
        return None
    for fmt in ("%d-%m-%Y", "%d/%m/%Y", "%d-%m-%y", "%d/%m/%y"):
        try:
            return datetime.strptime(str(date_text), fmt)
        except Exception:
            pass
    return None


def _bill_no(invoice_no: str):
    inv = str(invoice_no or "").strip()
    return int(inv) if inv.isdigit() else inv


def _bill_sort_key(invoice_no: str):
    inv = str(invoice_no or "").strip()
    nums = re.findall(r"\d+", inv)
    if nums:
        return (0, int(nums[-1]), inv)
    return (1, 10**12, inv)


def _rate_value(rec: InvoiceRecord) -> float:
    taxable = float(rec.taxable_value or 0)
    tax = float(rec.cgst or 0) + float(rec.sgst or 0) + float(rec.igst or 0)
    if taxable > 0 and tax >= 0:
        return round((tax / taxable) * 100, 2)
    return 0.0


def _rate(rec: InvoiceRecord) -> str:
    val = _rate_value(rec)
    if not val:
        return "18%"
    # Display clean GST rates when tiny differences come from rounding or extraction correction.
    for standard in (0, 5, 12, 18, 28):
        if abs(val - standard) <= 0.35:
            return f"{standard}%"
    return f"{int(val) if val.is_integer() else val}%"


def _is_special_customer(rec: InvoiceRecord, config: dict) -> bool:
    gstin = str(rec.gstin or "").strip().upper()
    name = str(rec.party_name or "").strip().lower()
    special_gstins = {str(x).strip().upper() for x in config.get("special_customer_gstins", [])}
    special_names = [str(x).strip().lower() for x in config.get("special_customer_names", []) if str(x).strip()]
    if gstin and gstin in special_gstins:
        return True
    if name and any(x in name for x in special_names):
        return True
    # Auto-detect lower GST customers: if tax rate is clearly below normal 18%.
    normal_rate = float(config.get("normal_tax_rate", 18) or 18)
    rate = _rate_value(rec)
    return bool(rate and rate < normal_rate - 0.5)


def _first_hsn(items: List[ItemRecord], rec: InvoiceRecord) -> str:
    if getattr(rec, "hsn", ""):
        return str(rec.hsn)
    for item in items:
        if str(item.invoice_no) == str(rec.invoice_no) and item.hsn_sac:
            return item.hsn_sac
    return "998346"


def _apply_border(ws, min_row: int = 1, max_col: int = 11):
    thin = Side(style="thin", color=BORDER_COLOR)
    for row in ws.iter_rows(min_row=min_row, max_row=ws.max_row, min_col=1, max_col=max_col):
        for cell in row:
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
            cell.alignment = Alignment(vertical="center", wrap_text=False)


def _setup_sheet(ws, title: str, business_name: str, business_gstin: str, month_name: str, year: str):
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A4"

    ws.merge_cells("A1:B1")
    ws["A1"] = title.upper()
    ws["C1"] = f"GSTIN- {business_gstin}" if business_gstin else "GSTIN-"
    ws.merge_cells("D1:K1")
    ws["D1"] = business_name.upper() if business_name else "DEMO TEXTILE RESEARCH LAB"
    ws["A2"] = "MONTH"
    ws["B2"] = month_name.upper()
    ws["C2"] = year

    for row in (1, 2):
        for col in range(1, 12):
            ws.cell(row, col).fill = PatternFill("solid", fgColor=LIGHT_BLUE)
            ws.cell(row, col).font = Font(bold=True, size=11)

    for col, header in enumerate(HEADERS, start=1):
        cell = ws.cell(3, col, header)
        cell.font = Font(bold=True, color="000000")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.fill = PatternFill("solid", fgColor=HEADER_BLUE)

    widths = {"A": 13, "B": 9, "C": 38, "D": 18, "E": 10, "F": 12, "G": 12, "H": 12, "I": 12, "J": 13, "K": 8}
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    ws.row_dimensions[1].height = 22
    ws.row_dimensions[3].height = 22


def _write_month_sheet(wb: Workbook, sheet_name: str, records: List[InvoiceRecord], items: List[ItemRecord], config: dict):
    ws = wb.create_sheet(_safe_sheet_name(sheet_name))
    d = _parse_date(records[0].invoice_date) if records else None
    month_name = MONTH_FULL[d.month - 1] if d else sheet_name.split()[0]
    year = str(d.year) if d else ""
    title = str(config.get("default_record_type", "Sales") or "Sales")
    business_name = config.get("business_name", "Demo Textile Research Lab")
    business_gstin = config.get("business_gstin", "00ABCDE1234A1Z0")
    _setup_sheet(ws, title, business_name, business_gstin, month_name, year)

    records = sorted(records, key=lambda r: _bill_sort_key(r.invoice_no))
    row_idx = 4
    for rec in records:
        dt = _parse_date(rec.invoice_date)
        ws.cell(row_idx, 1, dt if dt else rec.invoice_date)
        ws.cell(row_idx, 2, _bill_no(rec.invoice_no))
        ws.cell(row_idx, 3, str(rec.party_name or "").replace(" ,", "").strip())
        ws.cell(row_idx, 4, rec.gstin or "")
        ws.cell(row_idx, 5, _first_hsn(items, rec))
        ws.cell(row_idx, 6, rec.taxable_value or 0)
        ws.cell(row_idx, 7, rec.igst if rec.igst else "")
        ws.cell(row_idx, 8, rec.cgst if rec.cgst else "")
        ws.cell(row_idx, 9, rec.sgst if rec.sgst else "")
        ws.cell(row_idx, 10, rec.total or 0)
        ws.cell(row_idx, 11, _rate(rec))
        if _is_special_customer(rec, config):
            fill = PatternFill("solid", fgColor=SPECIAL_ORANGE)
            for c in range(1, 12):
                ws.cell(row_idx, c).fill = fill
                ws.cell(row_idx, c).font = Font(bold=True, color=SPECIAL_RED)
        elif rec.status == "CHECK":
            fill = PatternFill("solid", fgColor=YELLOW)
            for c in range(1, 12):
                ws.cell(row_idx, c).fill = fill
        elif rec.status == "ERROR":
            fill = PatternFill("solid", fgColor=RED)
            for c in range(1, 12):
                ws.cell(row_idx, c).fill = fill
        row_idx += 1

    total_row = row_idx + 1
    ws.cell(total_row, 5, "TOTAL")
    for col in range(6, 11):
        letter = get_column_letter(col)
        ws.cell(total_row, col, f"=SUM({letter}4:{letter}{row_idx-1})")
        ws.cell(total_row, col).number_format = '0.00'
    for col in range(1, 12):
        ws.cell(total_row, col).fill = PatternFill("solid", fgColor=GREEN)
        ws.cell(total_row, col).font = Font(bold=True)
        ws.cell(total_row, col).alignment = Alignment(horizontal="center" if col in (5, 11) else "right")

    _apply_border(ws, max_col=11)
    for row in ws.iter_rows(min_row=4, max_row=ws.max_row, min_col=6, max_col=10):
        for cell in row:
            cell.number_format = '0.00'
            cell.alignment = Alignment(horizontal="right", vertical="center")
    for cell in ws.iter_rows(min_row=4, max_row=ws.max_row, min_col=1, max_col=1):
        for x in cell:
            x.number_format = 'dd-mm-yy'
    for row in ws.iter_rows(min_row=4, max_row=ws.max_row, min_col=2, max_col=5):
        for cell in row:
            cell.alignment = Alignment(horizontal="center" if cell.column in (2, 4, 5) else "left", vertical="center")
    ws.auto_filter.ref = f"A3:K{max(ws.max_row,3)}"
    return ws


def _write_gstin_all(wb: Workbook, records: List[InvoiceRecord]):
    ws = wb.create_sheet("GSTIN ALL")
    headers = ["GSTIN", "NAME"]
    for col, h in enumerate(headers, 1):
        ws.cell(1, col, h)
        ws.cell(1, col).font = Font(bold=True, color=WHITE)
        ws.cell(1, col).fill = PatternFill("solid", fgColor=NAVY)
        ws.cell(1, col).alignment = Alignment(horizontal="center")
    seen: Dict[str, str] = {}
    for r in records:
        if r.gstin and r.party_name:
            seen.setdefault(r.gstin, r.party_name)
    for row_idx, (gstin, name) in enumerate(sorted(seen.items(), key=lambda x: x[1]), 2):
        ws.cell(row_idx, 1, gstin)
        ws.cell(row_idx, 2, name)
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 42
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:B{max(ws.max_row,1)}"
    _apply_border(ws, max_col=2)


def _sequence_warnings(records: List[InvoiceRecord]) -> List[str]:
    nums = []
    for r in records:
        inv = str(r.invoice_no or "")
        found = re.findall(r"\d+", inv)
        if found:
            nums.append(int(found[-1]))
    if not nums:
        return []
    warnings = []
    unique = sorted(set(nums))
    missing = [n for n in range(unique[0], unique[-1] + 1) if n not in set(unique)]
    if missing:
        warnings.append("Missing bill sequence numbers: " + ", ".join(map(str, missing[:50])) + ("..." if len(missing) > 50 else ""))
    if len(nums) != len(unique):
        warnings.append("Duplicate bill numbers detected")
    return warnings


def _write_summary(wb: Workbook, records: List[InvoiceRecord], clean: bool = False):
    ws = wb.create_sheet("SUMMARY", 0)
    ws.sheet_view.showGridLines = False
    ws.merge_cells("A1:D1")
    ws["A1"] = "GST REGISTER SUMMARY"
    ws["A1"].font = Font(bold=True, size=16, color=WHITE)
    ws["A1"].fill = PatternFill("solid", fgColor=NAVY)
    ws["A1"].alignment = Alignment(horizontal="center")

    ok = [r for r in records if r.status != "ERROR"]
    data = [
        ("Total Files Processed", len(records)),
        ("Records Exported", len(ok)),
        ("Needs Checking", sum(1 for r in records if r.status == "CHECK")),
        ("Errors", sum(1 for r in records if r.status == "ERROR")),
        ("Taxable", sum(r.taxable_value or 0 for r in ok)),
        ("IGST", sum(r.igst or 0 for r in ok)),
        ("CGST", sum(r.cgst or 0 for r in ok)),
        ("SGST", sum(r.sgst or 0 for r in ok)),
        ("Total", sum(r.total or 0 for r in ok)),
    ]
    ws.append([])
    ws.append(["Metric", "Value"])
    for k, v in data:
        ws.append([k, v])
    if not clean:
        warn_start = ws.max_row + 2
        ws.cell(warn_start, 1, "Validation Notes")
        ws.cell(warn_start, 1).font = Font(bold=True)
        warnings = _sequence_warnings(ok)
        if not warnings and not any(r.status in ("CHECK", "ERROR") for r in records):
            warnings = ["No major validation warnings found."]
        for w in warnings:
            ws.append([w])

    for cell in ws[3]:
        cell.font = Font(bold=True, color=WHITE)
        cell.fill = PatternFill("solid", fgColor=NAVY)
    for row in range(8, 12):
        ws.cell(row, 2).number_format = '0.00'
    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 18
    _apply_border(ws, max_col=4)


def _write_errors(wb: Workbook, records: List[InvoiceRecord]):
    bad = [r for r in records if r.status in ("CHECK", "ERROR", "DUPLICATE")]
    ws = wb.create_sheet("CHECK ERRORS")
    headers = ["STATUS", "NOTES", "SOURCE FILE", "BILL NO", "DATE", "NAME", "GSTIN", "TAXABLE", "FREIGHT", "IGST", "CGST", "SGST", "TOTAL"]
    for col, h in enumerate(headers, 1):
        ws.cell(1, col, h)
        ws.cell(1, col).font = Font(bold=True, color=WHITE)
        ws.cell(1, col).fill = PatternFill("solid", fgColor=NAVY)
        ws.cell(1, col).alignment = Alignment(horizontal="center")
    if not bad:
        ws.append(["OK", "No issues found"])
    else:
        for r_i, r in enumerate(bad, 2):
            vals = [r.status, r.notes, r.source_file, r.invoice_no, r.invoice_date, r.party_name, r.gstin, r.taxable_value or 0, r.freight or 0, r.igst or 0, r.cgst or 0, r.sgst or 0, r.total or 0]
            for c_i, v in enumerate(vals, 1):
                ws.cell(r_i, c_i, v)
    for col in range(8, 14):
        for row in range(2, max(ws.max_row, 2) + 1):
            ws.cell(row, col).number_format = '0.00'
    widths = {"A":12,"B":55,"C":50,"D":12,"E":12,"F":38,"G":18,"H":12,"I":12,"J":12,"K":12,"L":12,"M":12}
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:M{max(ws.max_row,1)}"
    _apply_border(ws, max_col=13)


def _write_audit(wb: Workbook, records: List[InvoiceRecord]):
    ws = wb.create_sheet("AUDIT")
    headers = ["BILL NO", "FILE", "STATUS", "CONFIDENCE", "NOTES", "FREIGHT INCLUDED", "SPECIAL/LOW GST"]
    for c, h in enumerate(headers, 1):
        ws.cell(1, c, h)
        ws.cell(1, c).font = Font(bold=True, color=WHITE)
        ws.cell(1, c).fill = PatternFill("solid", fgColor=NAVY)
    for i, r in enumerate(sorted(records, key=lambda x: _bill_sort_key(x.invoice_no)), 2):
        ws.cell(i, 1, r.invoice_no)
        ws.cell(i, 2, r.source_file)
        ws.cell(i, 3, r.status)
        ws.cell(i, 4, r.confidence)
        ws.cell(i, 5, r.notes)
        ws.cell(i, 6, r.freight or "")
        ws.cell(i, 7, "YES" if _is_special_customer(r, {"normal_tax_rate": 18}) else "")
    widths = {"A":10,"B":50,"C":12,"D":12,"E":65,"F":15,"G":18}
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:G{max(ws.max_row,1)}"
    _apply_border(ws, max_col=7)



def _build_workbook(records: List[InvoiceRecord], items: List[ItemRecord], *, clean_upload: bool = False) -> Workbook:
    wb = Workbook()
    wb.remove(wb.active)
    config = {
        "business_name": "Demo Textile Research Lab",
        "business_gstin": "00ABCDE1234A1Z0",
        "default_record_type": records[0].record_type if records else "Sales",
        "normal_tax_rate": 18,
        # Add exact GSTINs/names here later if needed. Low-GST rows are also auto-detected.
        "special_customer_gstins": [],
        "special_customer_names": [],
    }

    records = sorted(records, key=lambda r: _bill_sort_key(r.invoice_no))
    _write_summary(wb, records, clean=clean_upload)

    valid_records = [r for r in records if r.status != "ERROR"]
    grouped = defaultdict(list)
    for rec in valid_records:
        d = _parse_date(rec.invoice_date)
        if d:
            grouped[(d.year, d.month)].append(rec)
        else:
            grouped[(9999, 99)].append(rec)

    for (year, month), recs in sorted(grouped.items()):
        if month == 99:
            sheet_name = "UNKNOWN DATE"
        else:
            sheet_name = f"{MONTH_NAMES[month-1]}. {year}"
        _write_month_sheet(wb, sheet_name, recs, items, config)

    if not clean_upload:
        _write_gstin_all(wb, valid_records)
        _write_errors(wb, records)
        _write_audit(wb, records)
    return wb


def save_gst_workbook(output_path: str | Path, records: List[InvoiceRecord], items: List[ItemRecord], create_clean_upload: bool = True) -> tuple[Path, Path | None]:
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    master_wb = _build_workbook(records, items, clean_upload=False)
    master_wb.save(output_path)

    clean_path = None
    if create_clean_upload:
        clean_path = output_path.with_name(output_path.stem + "_READY_UPLOAD" + output_path.suffix)
        clean_wb = _build_workbook(records, items, clean_upload=True)
        clean_wb.save(clean_path)
    return output_path, clean_path
